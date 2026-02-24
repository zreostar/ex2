import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# 定義時間區間
TIME_SLOTS = [
    "08:00~08:50",
    "09:00~09:50",
    "10:00~10:50",
    "11:00~11:50",
    "01:00~01:50",
    "02:00~02:50",
    "03:00~03:50",
    "04:00~04:50"
]
DAYS = ["週一", "週二", "週三", "週四", "週五"]

# 課程資料
COURSES = [
    {"name": "電子實習", "teacher": "邱聰輝教授", "day": "週一", "start": "08:00", "end": "10:50"},
    {"name": "雲端化無限存取實務", "teacher": "劉恩成教授", "day": "週二", "start": "13:00", "end": "15:50"},
    {"name": "工智慧概論", "teacher": "賴文政教授", "day": "週三", "start": "08:00", "end": "08:50"},
    {"name": "工程數學", "teacher": "郭慶祥教授", "day": "週三", "start": "09:00", "end": "11:00"},
    {"name": "專題製作", "teacher": "祁存廣教授", "day": "週三", "start": "15:00", "end": "16:50"},
    {"name": "英文聽講", "teacher": "吳柏德教授", "day": "週四", "start": "08:00", "end": "09:50"},
    {"name": "自動控制與實習", "teacher": "謝飛虎教授", "day": "週四", "start": "13:00", "end": "16:50"},
    {"name": "工程數學", "teacher": "郭慶祥教授", "day": "週五", "start": "08:00", "end": "08:50"},
    {"name": "人工智慧概論", "teacher": "賴文政教授", "day": "週五", "start": "09:00", "end": "10:50"},
    {"name": "影像處理概論", "teacher": "王柏仁教授", "day": "週五", "start": "13:00", "end": "15:50"},
]

# 時間區間對應
slot_time_map = {
    "08:00~08:50": (8, 0, 8, 50),
    "09:00~09:50": (9, 0, 9, 50),
    "10:00~10:50": (10, 0, 10, 50),
    "11:00~11:50": (11, 0, 11, 50),
    "01:00~01:50": (13, 0, 13, 50),
    "02:00~02:50": (14, 0, 14, 50),
    "03:00~03:50": (15, 0, 15, 50),
    "04:00~04:50": (16, 0, 16, 50),
}

def time_to_minutes(hhmm):
    h, m = map(int, hhmm.split(":"))
    return h * 60 + m

def slot_range(start, end):
    start_min = time_to_minutes(start)
    end_min = time_to_minutes(end)
    slots = []
    for slot in TIME_SLOTS:
        s_h, s_m, e_h, e_m = slot_time_map[slot]
        slot_start = s_h * 60 + s_m
        slot_end = e_h * 60 + e_m
        if slot_start >= start_min and slot_end <= end_min:
            slots.append(slot)
    return slots

def build_timetable():
    # timetable[day][slot] = (course_name, teacher)
    timetable = {day: {slot: None for slot in TIME_SLOTS} for day in DAYS}
    merge_info = {day: {} for day in DAYS}
    conflicts = []
    for course in COURSES:
        day = course["day"]
        slots = slot_range(course["start"], course["end"])
        content = f"{course['name']}\n/ {course['teacher']}"
        for slot in slots:
            if timetable[day][slot] is not None:
                conflicts.append((day, slot))
            timetable[day][slot] = content
        merge_info[day][slots[0]] = len(slots)
    return timetable, merge_info, conflicts

def create_excel(timetable, merge_info, conflicts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "課表"
    # 欄位標題
    ws.cell(row=1, column=1, value="時間")
    for i, day in enumerate(DAYS):
        ws.cell(row=1, column=i+2, value=day)
    # 時間欄
    for i, slot in enumerate(TIME_SLOTS):
        ws.cell(row=i+2, column=1, value=slot)
    # 格式
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # 填入課程
    for col, day in enumerate(DAYS, start=2):
        row = 2
        while row <= len(TIME_SLOTS)+1:
            slot = TIME_SLOTS[row-2]
            val = timetable[day][slot]
            if val:
                merge_len = merge_info[day].get(slot, 1)
                ws.cell(row=row, column=col, value=val)
                if merge_len > 1:
                    ws.merge_cells(start_row=row, start_column=col, end_row=row+merge_len-1, end_column=col)
                # 格式
                for r in range(row, row+merge_len):
                    cell = ws.cell(row=r, column=col)
                    cell.alignment = align
                    cell.border = border
                    if (day, slot) in conflicts:
                        cell.fill = red_fill
                row += merge_len
            else:
                cell = ws.cell(row=row, column=col)
                cell.alignment = align
                cell.border = border
                row += 1
    # 衝突格子底色
    for day, slot in conflicts:
        col = DAYS.index(day)+2
        row = TIME_SLOTS.index(slot)+2
        ws.cell(row=row, column=col).fill = red_fill
    # 欄寬
    ws.column_dimensions[get_column_letter(1)].width = 13
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 22
    wb.save("timetable.xlsx")

def main():
    timetable, merge_info, conflicts = build_timetable()
    if conflicts:
        print("課表有衝突：")
        for day, slot in conflicts:
            print(f"{day} {slot}")
    create_excel(timetable, merge_info, conflicts)
    print("timetable.xlsx 已生成")

if __name__ == "__main__":
    main()
